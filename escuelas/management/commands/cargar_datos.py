# coding: utf-8
from __future__ import unicode_literals
import time
import datetime
import random, string
import hashlib

import progressbar
import requests
from openpyxl import load_workbook

from escuelas import models
from django.contrib.auth.models import User
from django.core.management.base import BaseCommand
from django.contrib.auth.models import Group
from django.contrib.auth.models import Permission
from django.contrib.contenttypes.models import ContentType

# Esta variable no se debe modificar. Se le puede cambiar el valor en tiempo
# de ejecución invocando el comando "make cargar_datos depuracion=1"
MODO_VERBOSE = False
PERFIL = False

def log(*k):
    global MODO_VERBOSE

    if MODO_VERBOSE:
        if isinstance(k, tuple):
            print " ".join(k)
        else:
            print k

BASE_URL = 'http://suite-api.dtelab.com.ar/api/'

def esperar(segundos):
    time.sleep(segundos)

def barra_de_progreso(simple=True):
    if simple:
        return progressbar.ProgressBar(widgets=[progressbar.SimpleProgress()])
    else:
        return progressbar.ProgressBar()

class Command(BaseCommand):
    help = 'Genera todos los datos iniciales.'

    def add_arguments(self, parser):
        parser.add_argument('--filtro', help="Aplica un filtro a los comandos que se ejecutaran")
        parser.add_argument('--depuracion', help="Permite activar todos los detalles (verbose mode)")
        parser.add_argument('--perfil_id', help="Permite aplicar un filtro de perfil al comando ejecutado")

    def handle(self, *args, **options):
        global MODO_VERBOSE
        global PERFIL
        filtro = options['filtro']
        depuracion = options['depuracion']
        perfil_id = options['perfil_id']

        print(u"perfil_id: " + perfil_id)


        # A continuación están todos los comandos que el importador
        # puede ejecutar. Cada comando es un método, con el mismo nombre
        # que aparece como cadena.
        #
        # por ejemplo 'crear_areas' es la orden para ejecutar el método
        # self.crear_areas().
        comandos = [
            'crear_cargos_escolares',
            'crear_regiones',
            'crear_tipos_de_financiamiento',
            'crear_niveles',
            'crear_modalidades',
            'crear_tipos_de_gestion',
            'crear_areas',
            'crear_programas',
            'crear_cargos',
            'crear_experiencias',
            'crear_contratos',
            'crear_motivos_de_tareas',
            'crear_estados_de_tareas',
            'crear_prioridades_de_tareas',
            'crear_estados_de_validaciones',
            'crear_motivos_de_conformaciones',
            'crear_categorias_de_eventos',
            'crear_estados_de_paquetes',

            'importar_distritos_y_localidades',
            'importar_escuelas',

            'limpiar_e_importar_permisos_con_grupos',
            'importar_usuarios',

            'importar_contactos',
            'importar_pisos',
            'vincular_programas',
            'importar_tareas',
            'importar_comentarios_de_tareas',
            'importar_eventos',
            'importar_eventos_por_perfil',
            'vincular_acompaniantes',
            'vincular_acompaniantes_por_perfil',
            'importar_conformaciones',
            'importar_validaciones',
            'importar_comentarios_de_validaciones',
            'importar_paquetes',
            'aplicar_permiso_sin_definir_a_los_perfiles_faltantes',
            'aplicar_password_inicial_para_usuarios',
            'aplicar_recibido_a_paquetes',
            'importar_estado_de_paquetes',
        ]


        print("Procesando comandos a ejecutar ...")
        esperar(1)

        if perfil_id:
            PERFIL = perfil_id
            print(PERFIL)
            print(u"Se ejecutarán los comandos solo para el perfil %s") %(perfil_id)

        if depuracion != '0':
            print(u"Modo depuracion activado.")
            MODO_VERBOSE = True
        else:
            print(u"Modo depuracion desactivado.")
            MODO_VERBOSE = False

        if filtro:
            print("Aplicando el filtro: " + filtro)
            comandos_filtrados = [x for x in comandos if filtro in x]

            cantidad_de_comandos = len(comandos_filtrados)

            if cantidad_de_comandos == 0:
                print("No hay ningún comando que coincida con el filtro.")
            else:
                print("Solo se ejecutaran %d comandos (de los %d disponibles)" %(cantidad_de_comandos, len(comandos)))
                self.ejecutar_comandos(comandos_filtrados)
        else:
            print("Se ejecutaran %d comandos" %(len(comandos)))
            self.ejecutar_comandos(comandos)


    def ejecutar_comandos(self, comandos):
        esperar(1)
        self.listar_comandos(comandos)
        esperar(2)

        for i, x in enumerate(comandos):
            print("[01;32m[%d/%d] Ejecutando tarea: %s [0m" %(i+1, len(comandos)-i, x))
            metodo = getattr(self, x)
            metodo()

    def listar_comandos(self, comandos):
        print("Se ejecutaran los comandos en este orden:")
        for i, x in enumerate(comandos):
            print ("  %d %s" %(i+1, x))

    def crear_regiones(self):
        numeros = range(1, 26)

        print("Creando Regiones")
        bar = barra_de_progreso()

        for n in bar(numeros):
            p, created = models.Region.objects.get_or_create(numero=n)
            log(str(p))

        p, created = models.Region.objects.get_or_create(numero=27)

        if MODO_VERBOSE:
            print(p)

    def importar_distritos_y_localidades(self):
        localidades = self.obtener_datos_desde_api('localidades')['localidades']

        print("Creando Localidades")
        bar = barra_de_progreso(simple=False)

        for localidad in bar(localidades):
            objeto_distrito, Localidadescreated = models.Distrito.objects.get_or_create(nombre=localidad['distrito'].title())
            objeto_localidad, created = models.Localidad.objects.get_or_create(nombre=localidad['localidad'].title())

            objeto_localidad.distrito = objeto_distrito
            objeto_localidad.save()

            objeto_distrito.region, created = models.Region.objects.get_or_create(numero=int(localidad['region']))
            objeto_distrito.save()

            if MODO_VERBOSE:
                print objeto_distrito, " -> ", objeto_localidad, "de la", objeto_distrito.region

    def importar_eventos(self):
        eventos = self.obtener_datos_desde_api('eventos')['eventos']
        cantidad_de_eventos_creados = 0
        cantidad_de_eventos_omitidos = 0
        cantidad_de_eventos_omitidos_por_perfil = 0
        cantidad_de_eventos_omitidos_por_escuela = 0
        cantidad_de_eventos_omitidos_por_categoria = 0

        print("Creando Eventos")
        bar = barra_de_progreso(simple=False)



        for evento in bar(eventos):
            legacy_id = evento['legacy_id']
            fecha_inicio = evento['fecha_inicio']
            hora_inicio = evento['hora_inicio']
            fecha_final = evento['fecha_final']
            hora_final = evento['hora_final']
            fecha_carga = evento['fecha_de_carga']
            cue = evento['cue']
            #responsable = evento['usuario']
            dni_usuario = evento['dni_usuario']
            objetivo = evento['objetivo']
            cantidad_de_participantes = evento['cantidad_de_participantes']
            minuta = evento['minuta']
            acta = evento['acta']
            categoria = evento['categoria'].capitalize()
            subcategoria = evento['subcategoria'].capitalize()

            if categoria == "Asistencia a escuela":
                categoria = "Asistencia"

            if subcategoria == "Servicio técnico de netbook":
                subcategoria = "Servicio técnico a netbook"

            titulo = categoria + " " + subcategoria

            categoria_2 = categoria + "/" + subcategoria

            if MODO_VERBOSE:
                print "========================================================================================================================"
                print "   Se intenta crear el evento con legacy_id: " + str(legacy_id) + ", asociado a cue: " + str(cue)
                print "========================================================================================================================"
                print "legacy_id:               " + str(legacy_id)
                print "Titulo:                  " + titulo
                print "Inicio:                  " + fecha_inicio + " " + hora_inicio
                print "Fin:                     " + fecha_final + " " + hora_final
                print "Categoria:               " + categoria
                print "Subcategoria:            " + subcategoria
                print "Categoria 2:             " + categoria_2
                print "Objetivo:                " + objetivo
                print "Acta:                    " + acta
                print "Responsable:             " + "dni " + dni_usuario
                print "Fecha de creacion:       " + fecha_carga
                print "Cant. de Participantes:  " + str(cantidad_de_participantes)
                print "Escuela:                 " + str(cue)
                print "=============================="


            try:
                objeto_responsable = models.Perfil.objects.get(dni=dni_usuario)
            except models.Perfil.DoesNotExist:
                log("Error, no existe registro de usuario buscado %s. No se registrará el evento." %(dni_usuario))
                cantidad_de_eventos_omitidos += 1
                cantidad_de_eventos_omitidos_por_perfil += 1
                continue

            try:
                objeto_escuela = models.Escuela.objects.get(cue=cue)
            except models.Escuela.DoesNotExist:
                log("Error, no existe la escuela buscada con cue %s. No se registrará el evento." %(cue))
                cantidad_de_eventos_omitidos += 1
                cantidad_de_eventos_omitidos_por_escuela += 1
                continue

            try:
                objeto_categoria = models.CategoriaDeEvento.objects.get(nombre=categoria_2)
            except models.CategoriaDeEvento.DoesNotExist:
                log("Error, no existe la categoria %s. No se registrará el evento" %(categoria_2))
                cantidad_de_eventos_omitidos += 1
                cantidad_de_eventos_omitidos_por_categoria += 1
                continue

            try:
                objeto_evento = models.Evento.objects.get(legacy_id=legacy_id)
            except models.Evento.DoesNotExist:
                objeto_evento = models.Evento(legacy_id=legacy_id)

            objeto_evento.responsable = objeto_responsable
            objeto_evento.escuela = objeto_escuela
            objeto_evento.categoria = objeto_categoria
            objeto_evento.titulo = titulo
            objeto_evento.fecha = fecha_inicio
            objeto_evento.inicio = hora_inicio
            objeto_evento.fecha_fin = fecha_final
            objeto_evento.fin = hora_final
            objeto_evento.objetivo = objetivo
            objeto_evento.cantidadDeParticipantes = cantidad_de_participantes
            objeto_evento.minuta = minuta
            objeto_evento.acta_legacy = acta

            objeto_evento.save()
            cantidad_de_eventos_creados += 1


            if MODO_VERBOSE:
                print "=============================="
                print "   SE HA CREADO EL REGISTRO   "
                print "=============================="
                print "legacy_id:               " + str(legacy_id)
                print "Titulo:                  " + titulo
                print "Inicio:                  " + fecha_inicio + " " + hora_inicio
                print "Fin:                     " + fecha_final + " " + hora_final
                print "Categoria:               " + categoria_2
                print "Objetivo:                " + objetivo
                print "Acta:                    " + acta
                print "Responsable:             " + objeto_responsable.apellido + ", " + objeto_responsable.nombre + " (dni " + dni_usuario + ")"
                print "Fecha de creacion:       " + fecha_carga
                print "Cant. de Participantes:  " + str(cantidad_de_participantes)
                print "Escuela:                 " + objeto_escuela.nombre + " " + str(cue)
                print "=============================="

        print("   Se crearon %d eventos correctamente." %(cantidad_de_eventos_creados))
        print "Total de eventos omitidos: %d" %(cantidad_de_eventos_omitidos)
        print "Eventos omitidos por no encontrar escuela: %d" %(cantidad_de_eventos_omitidos_por_escuela)
        print "Eventos omitidos por no encontrar perfil: %d" %(cantidad_de_eventos_omitidos_por_perfil)
        print "Eventos omitidos por no encontrar categoria: %d" %(cantidad_de_eventos_omitidos_por_categoria)

    def importar_eventos_por_perfil(self):
        if PERFIL:
            print (u"El ID de perfil es " + PERFIL)
        ruta = "eventos_por_perfil?perfil_id=" + PERFIL
        eventos = self.obtener_datos_desde_api(ruta)['eventos_por_perfil']
        cantidad_de_eventos_creados = 0
        cantidad_de_eventos_omitidos = 0
        cantidad_de_eventos_omitidos_por_perfil = 0
        cantidad_de_eventos_omitidos_por_escuela = 0
        cantidad_de_eventos_omitidos_por_categoria = 0

        print("Creando Eventos")
        bar = barra_de_progreso(simple=False)



        for evento in bar(eventos):
            legacy_id = evento['legacy_id']
            fecha_inicio = evento['fecha_inicio']
            hora_inicio = evento['hora_inicio']
            fecha_final = evento['fecha_final']
            hora_final = evento['hora_final']
            fecha_carga = evento['fecha_de_carga']
            cue = evento['cue']
            #responsable = evento['usuario']
            dni_usuario = evento['dni_usuario']
            objetivo = evento['objetivo']
            cantidad_de_participantes = evento['cantidad_de_participantes']
            minuta = evento['minuta']
            acta = evento['acta']
            categoria = evento['categoria'].capitalize()
            subcategoria = evento['subcategoria'].capitalize()

            if categoria == "Asistencia a escuela":
                categoria = "Asistencia"

            if subcategoria == "Servicio técnico de netbook":
                subcategoria = "Servicio técnico a netbook"

            titulo = categoria + " " + subcategoria

            categoria_2 = categoria + "/" + subcategoria

            if MODO_VERBOSE:
                print "========================================================================================================================"
                print "   Se intenta crear el evento con legacy_id: " + str(legacy_id) + ", asociado a cue: " + str(cue)
                print "========================================================================================================================"
                print "legacy_id:               " + str(legacy_id)
                print "Titulo:                  " + titulo
                print "Inicio:                  " + fecha_inicio + " " + hora_inicio
                print "Fin:                     " + fecha_final + " " + hora_final
                print "Categoria:               " + categoria
                print "Subcategoria:            " + subcategoria
                print "Categoria 2:             " + categoria_2
                print "Objetivo:                " + objetivo
                print "Acta:                    " + acta
                print "Responsable:             " + "dni " + dni_usuario
                print "Fecha de creacion:       " + fecha_carga
                print "Cant. de Participantes:  " + str(cantidad_de_participantes)
                print "Escuela:                 " + str(cue)
                print "=============================="


            try:
                objeto_responsable = models.Perfil.objects.get(dni=dni_usuario)
            except models.Perfil.DoesNotExist:
                log("Error, no existe registro de usuario buscado %s. No se registrará el evento." %(dni_usuario))
                cantidad_de_eventos_omitidos += 1
                cantidad_de_eventos_omitidos_por_perfil += 1
                continue

            try:
                objeto_escuela = models.Escuela.objects.get(cue=cue)
            except models.Escuela.DoesNotExist:
                log("Error, no existe la escuela buscada con cue %s. No se registrará el evento." %(cue))
                cantidad_de_eventos_omitidos += 1
                cantidad_de_eventos_omitidos_por_escuela += 1
                continue

            try:
                objeto_categoria = models.CategoriaDeEvento.objects.get(nombre=categoria_2)
            except models.CategoriaDeEvento.DoesNotExist:
                log("Error, no existe la categoria %s. No se registrará el evento" %(categoria_2))
                cantidad_de_eventos_omitidos += 1
                cantidad_de_eventos_omitidos_por_categoria += 1
                continue

            try:
                objeto_evento = models.Evento.objects.get(legacy_id=legacy_id)
            except models.Evento.DoesNotExist:
                objeto_evento = models.Evento(legacy_id=legacy_id)

            objeto_evento.responsable = objeto_responsable
            objeto_evento.escuela = objeto_escuela
            objeto_evento.categoria = objeto_categoria
            objeto_evento.titulo = titulo
            objeto_evento.fecha = fecha_inicio
            objeto_evento.inicio = hora_inicio
            objeto_evento.fecha_fin = fecha_final
            objeto_evento.fin = hora_final
            objeto_evento.objetivo = objetivo
            objeto_evento.cantidadDeParticipantes = cantidad_de_participantes
            objeto_evento.minuta = minuta
            objeto_evento.acta_legacy = acta

            objeto_evento.save()
            cantidad_de_eventos_creados += 1


            if MODO_VERBOSE:
                print "=============================="
                print "   SE HA CREADO EL REGISTRO   "
                print "=============================="
                print "legacy_id:               " + str(legacy_id)
                print "Titulo:                  " + titulo
                print "Inicio:                  " + fecha_inicio + " " + hora_inicio
                print "Fin:                     " + fecha_final + " " + hora_final
                print "Categoria:               " + categoria_2
                print "Objetivo:                " + objetivo
                print "Acta:                    " + acta
                print "Responsable:             " + objeto_responsable.apellido + ", " + objeto_responsable.nombre + " (dni " + dni_usuario + ")"
                print "Fecha de creacion:       " + fecha_carga
                print "Cant. de Participantes:  " + str(cantidad_de_participantes)
                print "Escuela:                 " + objeto_escuela.nombre + " " + str(cue)
                print "=============================="

        print("   Se crearon %d eventos correctamente." %(cantidad_de_eventos_creados))
        print "Total de eventos omitidos: %d" %(cantidad_de_eventos_omitidos)
        print "Eventos omitidos por no encontrar escuela: %d" %(cantidad_de_eventos_omitidos_por_escuela)
        print "Eventos omitidos por no encontrar perfil: %d" %(cantidad_de_eventos_omitidos_por_perfil)
        print "Eventos omitidos por no encontrar categoria: %d" %(cantidad_de_eventos_omitidos_por_categoria)


    def importar_escuelas(self):
        resultado = self.obtener_datos_desde_api('escuelas')

        print("Se importarán %d escuelas en total." %(resultado['cantidad']))
        esperar(2)

        escuelas = resultado['escuelas']

        bar = barra_de_progreso(simple=False)

        for escuela in bar(escuelas):

            if MODO_VERBOSE:
                print "Intentando crear el registro escuela id_original:", escuela['id']

            objeto_escuela, created = models.Escuela.objects.get_or_create(cue=escuela['cue'])

            objeto_area, created = models.Area.objects.get_or_create(nombre=escuela['area'].title())

            objeto_localidad, created = models.Localidad.objects.get_or_create(nombre=escuela['localidad'].title())
            objeto_tipoDeFinanciamiento, created = models.TipoDeFinanciamiento.objects.get_or_create(nombre=escuela['tipo_financiamiento'].title())
            objeto_nivel, created = models.Nivel.objects.get_or_create(nombre=escuela['nivel'].title())
            objeto_tipoDeGestion, created = models.TipoDeGestion.objects.get_or_create(nombre=escuela['tipo_gestion'].title())
            #objeto_programa, created = models.Programa.objects.get_or_create(nombre=escuela['programa'].title())

            objeto_escuela.nombre = escuela['nombre'].title()
            objeto_escuela.direccion = escuela['direccion']
            objeto_escuela.telefono = escuela['telefono']
            objeto_escuela.latitud = escuela['latitud']
            objeto_escuela.longitud = escuela['longitud']

            objeto_escuela.area = objeto_area

            objeto_escuela.localidad = objeto_localidad
            objeto_escuela.tipoDeFinanciamiento = objeto_tipoDeFinanciamiento
            objeto_escuela.nivel = objeto_nivel
            objeto_escuela.tipoDeGestion = objeto_tipoDeGestion
            #objeto_escuela.programas = objeto_programa

            objeto_escuela.estado = True

            objeto_escuela.save()

            log("Se ha creado el registro:")
            log(objeto_escuela, "\n CUE: ", objeto_escuela.cue, "\n Direccion: ", objeto_escuela.direccion, "\n Tel: ", objeto_escuela.telefono, "\n ", objeto_escuela.localidad, "\n ", objeto_escuela.area, "\n ", objeto_escuela.nivel, "\n ", objeto_escuela.tipoDeFinanciamiento, "\n ", objeto_escuela.tipoDeGestion)
            log("===========")

    def importar_usuarios(self):
        # ARCHIVO = './/archivos_para_importacion/dte_perfiles_2017.xlsx'
        ARCHIVO = './/archivos_para_importacion/dte_perfiles_09-2017.xlsx'
        LIMITE_DE_FILAS = 800

        print("Comenzando la importación de usuarios")
        log("Iniciando la importación del archivo: " + ARCHIVO)
        wb = load_workbook(ARCHIVO)

        columnas_como_string = ", ".join(wb.get_sheet_names())
        log("Las páginas de la planilla son: " + columnas_como_string)

        filas_procesadas = 0
        filas_omitidas_o_con_errores = 0
        filas_omitidas_lista = ""

        listado = ""

        def formatear_fecha(fecha):
            if fecha:
                return fecha.strftime('%Y-%m-%d')
            else:
                return fecha

        def obtener_valores_desde_fila(fila):
            return {
                "region":               fila[0].value,
                "cargo":                fila[1].value,
                "contrato":             fila[2].value,
                "carga_horaria":        fila[3].value,
                "consultor":            fila[4].value.strip().capitalize(),
                # "documentacion":        fila[5].value, # En principio, no nos interesa este dato.
                "expediente":           fila[6].value,
                "fechaDeRenuncia":      fila[7].value,
                "titulo":               fila[8].value,
                "fechaDeIngreso":       fila[9].value,
                "perfil":               fila[10].value,
                "dni":                  fila[11].value,
                "cuil":                 fila[12].value,
                "cbu":                  fila[13].value,
                "email":                fila[14].value,
                "email_laboral":        fila[15].value,
                "direccion":            fila[16].value,
                "localidad":            fila[17].value,
                "codigo_postal":        fila[18].value,
                "fechaDeNacimiento":    fila[19].value,
                "telefono_celular":     fila[20].value,
                "telefono_particular":  fila[21].value,
                "telefono_alternativo": fila[22].value,
                "rol": fila[23].value,
            }

        bar = barra_de_progreso(simple=False)
        #for conformacion in bar(conformaciones):

        cantidad_de_filas_con_datos = 0
        cantidad_de_filas_procesadas_sin_errores = 0

        for indice, fila in bar(enumerate(wb.active.rows)):

            if indice is 0:
                continue;             # Ignora la cabecera

            if not fila[1].value:
                log("Terminando en la fila %d porque no parece haber mas registros." %(indice + 1))
                break

            cantidad_de_filas_con_datos += 1
            log("Procesando fila '%d'" %(indice +1))

            try:
                valores = obtener_valores_desde_fila(fila)

                if valores['fechaDeRenuncia']:
                    log("  Perfil no activo (Renunció)")
                    fechaDeRenuncia=formatear_fecha(valores['fechaDeRenuncia'])
                else:
                    log("  Perfil activo")
                    fechaDeRenuncia=None

                region=str(valores['region'])

                if region.startswith('NC') or region.startswith('Nc') or region.startswith("ESP/NC"):
                    region="27"

                if valores['cargo']:
                    cargo=valores['cargo']
                else:
                    log("  No tiene cargo")

                contrato=valores['contrato']
                carga_horaria=valores['carga_horaria']
                consultor=valores['consultor'].split(',')
                apellido=consultor[0]
                nombre=consultor[1].title()

                if valores['expediente']:
                    expediente=valores['expediente']
                else:
                    log("  No tiene expediente")
                    expediente="Sin Datos"

                if valores['titulo']:
                    titulo=valores['titulo']
                else:
                    log("  No tiene título")
                    titulo="Sin Datos"

                fechaDeIngreso=formatear_fecha(valores['fechaDeIngreso'])

                if valores['perfil']:
                    experiencia=valores['perfil']
                else:
                    log("  No tiene perfil")
                    experiencia="Sin Datos"

                dni = str(valores['dni'])

                if valores['cuil']:
                    cuil=str(valores['cuil'])
                else:
                    log("  No tiene CUIL")
                    cuil="Sin Datos"

                if valores['cbu']:
                    cbu=valores['cbu']
                else:
                    log("  No tiene cbu")
                    cbu="Sin Datos"

                if valores['email']:
                    email=valores['email']
                else:
                    log("  No tiene email")
                    email="Sin Datos"

                if valores['email_laboral']:
                    email_laboral=valores['email_laboral']
                else:
                    log("  No tiene email laboral")
                    email_laboral=apellido+"@abc.gob.ar"

                email_laboral = email_laboral.lower()

                if valores['direccion']:
                    direccion=valores['direccion']
                else:
                    log("  No tiene direccion")
                    direccion="Sin Datos"

                localidad=valores['localidad'].title()
                codigo_postal=str(valores['codigo_postal'])

                if valores['fechaDeNacimiento']:
                    fechaDeNacimiento=formatear_fecha(valores['fechaDeNacimiento'])
                else:
                    log("  No tiene fecha de nacimiento")
                    fechaDeNacimiento=None

                if valores['telefono_celular']:
                    telefono_celular=valores['telefono_celular']
                else:
                    log("  No tiene telefono celular")
                    telefono_celular="Sin Datos"

                if valores['telefono_particular']:
                    telefono_particular=valores['telefono_particular']
                else:
                    log("  No tiene telefono Particular")
                    telefono_particular="Sin Datos"

                if valores['rol']:
                    grupo=valores['rol']
                else:
                    log("  No tiene ROL asignado")
                    grupo="Sin Definir"

                username=email_laboral
                default_pass="dte_"+dni

                try:
                    user = User.objects.get(username=email_laboral)
                except User.DoesNotExist:
                    user = User(username=email_laboral, email=email)
                    user.set_password(default_pass)

                user.save()

                perfil = models.Perfil.objects.get(user=user)

                try:
                    objeto_grupo = Group.objects.get(name=grupo)
                except Group.DoesNotExist:
                    log("  No existe el grupo", grupo)
                    continue

                perfil.group = objeto_grupo
                perfil.nombre = nombre
                perfil.apellido = apellido
                perfil.fechadenacimiento = fechaDeNacimiento
                perfil.titulo = titulo
                perfil.dni = dni
                perfil.cuit = cuil
                perfil.cbu = cbu
                perfil.email = email
                perfil.direccionCalle = direccion
                perfil.codigoPostal = codigo_postal
                perfil.telefonoCelular = telefono_celular
                perfil.telefonoAlternativo = telefono_particular
                perfil.expediente = expediente
                perfil.fechaDeIngreso = fechaDeIngreso
                perfil.fechaDeRenuncia = fechaDeRenuncia
                perfil.emailLaboral = email_laboral

                try:
                    perfil.region = models.Region.objects.get(numero=int(region))
                except models.Region.DoesNotExist:
                    log("  No existe la region consultada: %s" %(cargo))
                except ValueError:
                    log("  No existe la region consultada: %s" %(cargo))
                    continue

                # perfil.experiencia = models.Experiencia.objects.get(nombre=experiencia)
                # perfil.localidad = models.Localidad.objects.get(nombre=localidad)

                try:
                    objeto_cargo = models.Cargo.objects.get(nombre=cargo)
                except models.Cargo.DoesNotExist:
                    log("  No existe el cargo con ese nombre: %s" %(cargo))
                    continue

                perfil.cargo = objeto_cargo

                if contrato == "PLANTA" or contrato.lower() == 'planta':
                    contrato = "Planta"

                try:
                    perfil.contrato = models.Contrato.objects.get(nombre=contrato)
                except models.Contrato.DoesNotExist:
                    log("  No existe un contrato con ese nombre: %s" %(contrato))
                    continue

                perfil.save()
                cantidad_de_filas_procesadas_sin_errores += 1
            except TypeError, e:
                log("-----")
                log("Fila %d - ****** OMITIDA, TypeError. La fila contiene caracteres incorrectos." %(indice + 1))
                filas_omitidas_o_con_errores += 1
                filas_omitidas_lista += ", " + str(indice + 1)
                log(str(e))
                log("-----")
                continue

            log("Fila %d - Cargando datos de perfil para consultor: '%s'" %(indice + 1, valores["consultor"]))

            listado += apellido + nombre + "," + region + "," + email_laboral + "," + default_pass + "\n"


            filas_procesadas += 1

            if indice > LIMITE_DE_FILAS:
                break


        log("Terminó la ejecución")

        print("")
        print("Resumen:")
        print("")
        print(" - cantidad total de filas:                       " + str(cantidad_de_filas_con_datos))
        print(" - filas procesadas:                              " + str(cantidad_de_filas_procesadas_sin_errores))
        print(" - cantidad de filas que fallaron:                " + str(cantidad_de_filas_con_datos - cantidad_de_filas_procesadas_sin_errores))

        print("")
        print("")
        #print("Listado: ")
        #print("")
        #print listado
        f = open('listado_de_usuarios.csv', 'w')
        f.write(listado.encode('utf-8'))
        f.close()

    def aplicar_password_inicial_para_usuarios(self):
        usuarios = User.objects.order_by('username') # TODO: solo pedir los que no renunciaron.

        print("Nombre completo;region;usuario;password")

        for u in usuarios:

            if u.perfil.fechaDeRenuncia:
                continue

            if u.perfil.region:
                random_password = hashlib.md5(u.perfil.dni).hexdigest()[:8]
                u.set_password(random_password)
                u.save()
                print(u"%s, %s;%d;%s;%s" %(u.perfil.apellido, u.perfil.nombre, u.perfil.region.numero, u.username, random_password))


    def importar_conformaciones(self):
        resultado = self.obtener_datos_desde_api('conformaciones')

        print("Se importarán %d conformaciones en total." %(resultado['cantidad']))
        esperar(2)

        conformaciones = resultado['conformaciones']

        bar = barra_de_progreso(simple=False)

        for conformacion in bar(conformaciones):

            cue_conformado = conformacion['cue_conformado']
            cue_principal = conformacion['cue_principal']
            motivo = conformacion['motivo']
            fecha = conformacion['fecha']

            if MODO_VERBOSE:
                print "Intentando crear conformacion de escuela `", cue_conformado, "` con la escuela padre ", cue_principal, " y motivo ", motivo
                print "======================================================================================================"
                print "Escuela ", cue_conformado, " queda conformada con padre ", cue_principal
                print "Fecha:   ", fecha
                print "Motivo:  ", motivo
                print "======================================================================================================"

            objeto_escuela = models.Escuela.objects.get(cue=cue_conformado)
            escuela_padre = models.Escuela.objects.get(cue=cue_principal)
            objeto_motivo = models.MotivoDeConformacion.objects.get(nombre=motivo)
            fecha_como_objeto = datetime.datetime.strptime(fecha, "%Y-%m-%d")

            if objeto_escuela.padre != escuela_padre:
                escuela_padre.conformar_con(objeto_escuela, objeto_motivo, fecha_como_objeto)

            objeto_escuela.save()

    def importar_paquetes(self):
        resultado = self.obtener_datos_desde_api('paquetes')

        cantidad_de_paquetes_creados = 0
        cantidad_de_paquetes_omitidos = 0
        cantidad_de_paquetes_sin_escuela = 0

        print("Se importarán %d paquetes de provisión en total." %(resultado['cantidad']))
        esperar(2)

        paquetes = resultado['paquetes']

        bar = barra_de_progreso(simple=False)

        for paquete in bar(paquetes):

            legacy_id = paquete['legacy_id']
            cue = paquete['cue']
            ne = paquete['ne']
            servidor_serie = paquete['servidor_serie']
            idhardware = paquete['idhardware']
            marca_de_arranque = paquete['marca_de_arranque']
            llave_servidor = paquete['llave_servidor']
            fecha_pedido = paquete['fecha_pedido']
            comentario = paquete['comentario']
            carpeta_paquete = paquete['carpeta_paquete']
            fecha_envio_anses = paquete['fecha_envio_anses']
            zipanses = paquete['zipanses']
            estado = paquete['estado']
            fecha_devolucion = paquete['fecha_devolucion']
            id_devolucion = paquete['id_devolucion']
            leido = paquete['leido']


            if MODO_VERBOSE:
                print "======================================================================================================"
                print "Intentando crear paquete de provisión para el cue ", cue
                print "======================================================================================================"
                print "legacy_id", legacy_id
                print "CUE ", cue
                print "Fecha de pedido:", fecha_pedido
                print "NE:  ", ne
                print "Servidor:  ", servidor_serie
                print "idhardware:  ", idhardware
                print "marca_de_arranque:  ", marca_de_arranque
                print "Llave del servidor:  ", llave_servidor
                print "Comentario:  ", comentario
                print "Carpeta del paquete:  ", carpeta_paquete
                print "Fecha de envío a ANSES:  ", fecha_envio_anses
                print "ZIP ANSES:  ", zipanses
                print "Estado:  ", estado
                print "Fecha de devolución:  ", fecha_devolucion,
                print "ID devolución:  ", id_devolucion,
                print "Leido:  ", leido
                print "======================================================================================================"

            try:
                objeto_escuela = models.Escuela.objects.get(cue=cue)
            except models.Escuela.DoesNotExist:
                log("Error, no existe la escuela con cue %s. Se ignora el registro." %(cue))
                cantidad_de_paquetes_omitidos += 1
                cantidad_de_paquetes_sin_escuela += 1
                continue

            if estado == 0:
                estado = "Objetado"
            elif estado == 1:
                estado = "Pendiente"
            elif estado == 2:
                estado = "EducAr"
            elif estado == 3:
                estado = "Devuelto"
            else:
                estado = "Descargado"

            objeto_estado = models.EstadoDePaquete.objects.get(nombre=estado)

            if leido == 0:
                leido = False
            else:
                leido = True

            if fecha_envio_anses == "0000-00-00":
                fecha_envio_anses = None

            objeto_paquete, created = models.Paquete.objects.get_or_create(legacy_id=legacy_id)
            objeto_paquete.escuela = objeto_escuela
            objeto_paquete.fecha_pedido = fecha_pedido
            objeto_paquete.ne = ne
            objeto_paquete.id_hardware = idhardware
            objeto_paquete.marca_de_arranque = marca_de_arranque
            objeto_paquete.comentario = comentario
            objeto_paquete.carpeta_paquete = carpeta_paquete
            objeto_paquete.fecha_envio = fecha_envio_anses
            objeto_paquete.zip_paquete = zipanses
            objeto_paquete.estado = objeto_estado
            objeto_paquete.fecha_devolucion = fecha_devolucion
            objeto_paquete.id_devolucion = id_devolucion
            objeto_paquete.leido = leido

            objeto_paquete.save()

            cantidad_de_paquetes_creados += 1

        print("Resumen de paquetes:")
        print("   Se crearon %d paquetes correctamente." %(cantidad_de_paquetes_creados))
        print("   Se evitaron crear %d paquetes:" %(cantidad_de_paquetes_omitidos))
        print("     No se encontró la escuela de %s paquetes:" %(cantidad_de_paquetes_sin_escuela))


    def importar_validaciones(self):
        resultado = self.obtener_datos_desde_api('validaciones')
        cantidad_de_validaciones_creadas = 0
        cantidad_de_validaciones_omitidas = 0
        cantidad_de_validaciones_sin_escuela = 0
        cantidad_de_validaciones_sin_usuario = 0
        cantidad_de_validaciones_con_motivo_erroneo = 0

        print("Se importarán %d validaciones en total." %(resultado['cantidad']))
        esperar(2)

        validaciones = resultado['validaciones']

        bar = barra_de_progreso(simple=False)

        for validacion in bar(validaciones):

            legacy_id = validacion['legacy_id']
            cue = validacion['cue']
            escuela = validacion['escuela']
            dni_usuario = validacion['dni_usuario']
            usuario = validacion['usuario']
            estado = validacion['estado']
            fecha = validacion['fecha_de_alta']

            if estado==1:
                estado = "Pendiente"
            elif estado==2:
                estado = "Objetada"
            elif estado==3:
                estado = "Aprobada"
            else:
                estado = "NO IMPORTAR"

            if MODO_VERBOSE:
                print "Intentando crear validacion con legacy_id ", legacy_id, " y estado ", estado
                print "======================================================================================================"
                print "Escuela ", escuela, "  ", cue
                print "Usuario: ", usuario, " (", dni_usuario, ")"
                print "Fecha:   ", fecha
                print "Estado:  ", estado
                print "======================================================================================================"


            try:
                objeto_escuela = models.Escuela.objects.get(cue=cue)
            except models.Escuela.DoesNotExist:
                log("Error, no existe la escuela con cue %s. Se ignora el registro." %(cue))
                cantidad_de_validaciones_omitidas += 1
                cantidad_de_validaciones_sin_escuela += 1
                continue

            try:
                objeto_usuario = models.Perfil.objects.get(dni=dni_usuario)
            except models.Perfil.DoesNotExist:
                log("Error, no existe el usuario con dni %s. Se ignora el registro." %(dni_usuario))
                cantidad_de_validaciones_omitidas += 1
                cantidad_de_validaciones_sin_usuario += 1
                continue



            try:
                objeto_estado = models.EstadoDeValidacion.objects.get(nombre=estado)
            except models.EstadoDeValidacion.DoesNotExist:
                log("Error, estado %s no corresponde. Se ignora el registro." %(estado))
                cantidad_de_validaciones_omitidas += 1
                cantidad_de_validaciones_con_motivo_erroneo += 1
                continue


            objeto_validacion, created = models.Validacion.objects.get_or_create(legacy_id=legacy_id)
            objeto_validacion.fecha_de_alta = fecha
            objeto_validacion.autor = objeto_usuario
            objeto_validacion.estado = objeto_estado
            objeto_validacion.escuela = objeto_escuela

            objeto_validacion.save()

            cantidad_de_validaciones_creadas += 1

        print("Resumen de validaciones:")
        print("   Se crearon %d validaciones correctamente." %(cantidad_de_validaciones_creadas))
        print("   Se evitaron crear %d validaciones:" %(cantidad_de_validaciones_omitidas))
        print("     No se encontró la escuela de %s validaciones:" %(cantidad_de_validaciones_sin_escuela))
        print("     No se encontró el usuario de %s validaciones:" %(cantidad_de_validaciones_sin_usuario))
        print("     Hay %s validaciones con estado erróneo:" %(cantidad_de_validaciones_con_motivo_erroneo))

    def importar_comentarios_de_validaciones(self):
        resultado = self.obtener_datos_desde_api('historial_validaciones')
        cantidad_de_validaciones_creadas = 0
        cantidad_de_validaciones_omitidas = 0
        #cantidad_de_validaciones_sin_escuela = 0
        cantidad_de_validaciones_sin_usuario = 0
        cantidad_de_validaciones_con_motivo_erroneo = 0

        print("Se importarán %d comentarios de validaciones en total." %(resultado['cantidad']))
        esperar(2)

        comentarios = resultado['historial_validaciones']

        bar = barra_de_progreso(simple=False)

        for comentario in bar(comentarios):

            cantidad = comentario['cantidad']
            usuario = comentario['usuario']
            dni_usuario = comentario['dni_usuario']
            fecha = comentario['fecha']
            legacy_id = comentario['legacy_id']
            validacion = comentario['validacion_legacy_id']
            comentario = comentario['observaciones']

            if MODO_VERBOSE:
                print "Intentando crear comentario de validacion con legacy_id ", legacy_id
                print "======================================================================================================"
                print "ID Validacion:", validacion
                print "Usuario: ", usuario, " (", dni_usuario, ")"
                print "Fecha:   ", fecha
                print "Comentario:  ", comentario
                print "Cantidad:  ", cantidad
                print "======================================================================================================"


            try:
                objeto_usuario = models.Perfil.objects.get(dni=dni_usuario)
            except models.Perfil.DoesNotExist:
                log("Error, no existe el usuario con dni %s. Se ignora el registro." %(dni_usuario))
                cantidad_de_validaciones_omitidas += 1
                cantidad_de_validaciones_sin_usuario += 1
                continue

            try:
                objeto_validacion = models.Validacion.objects.get(legacy_id=validacion)
            except models.Validacion.DoesNotExist:
                log("Error, no existe la validacion con legacy_id %s. Se ignora el registro." %(validacion))
                cantidad_de_validaciones_omitidas += 1
                cantidad_de_validaciones_con_motivo_erroneo += 1
                continue


            objeto_comentario_de_validacion, created = models.ComentarioDeValidacion.objects.get_or_create(legacy_id=legacy_id)
            objeto_comentario_de_validacion.validacion = objeto_validacion
            objeto_comentario_de_validacion.fecha = fecha
            objeto_comentario_de_validacion.autor = objeto_usuario
            objeto_comentario_de_validacion.comentario = comentario
            objeto_comentario_de_validacion.cantidad = cantidad

            objeto_comentario_de_validacion.save()

            objeto_validacion.cantidad_pedidas = cantidad
            if (objeto_validacion.estado.nombre == "Aprobada"):
                objeto_validacion.cantidad_validadas = cantidad

            objeto_validacion.save()

            cantidad_de_validaciones_creadas += 1

        print("Resumen de comentarios de validaciones:")
        print("   Se crearon %d comentarios correctamente." %(cantidad_de_validaciones_creadas))
        print("   Se evitaron crear %d comentarios:" %(cantidad_de_validaciones_omitidas))
        print("     No se encontró la validacion correspondiente a %s comentarios:" %(cantidad_de_validaciones_con_motivo_erroneo))
        print("     No se encontró el usuario de %s comentarios:" %(cantidad_de_validaciones_sin_usuario))


    def importar_contactos(self):
        contactos = self.obtener_datos_desde_api('contactos')['contactos']

        print("Importando Contactos")
        bar = barra_de_progreso(simple=False)

        for contacto in bar(contactos):
            log("Buscando escuela para el contacto: ", contacto['escuela'])
            objeto_escuela = models.Escuela.objects.get(cue=contacto['escuela'])
            objeto_cargo = models.CargoEscolar.objects.get(nombre=contacto['cargo'])
            objeto_contacto, created = models.Contacto.objects.get_or_create(nombre=contacto['nombre'].title())
            #
            objeto_contacto.cargo = objeto_cargo
            objeto_contacto.escuela = objeto_escuela
            if contacto['email']:
                objeto_contacto.email = contacto['email'].lower()

            objeto_contacto.telefono_particular = contacto['telefono']
            objeto_contacto.telefono_celular = contacto['celular']
            if contacto['horario']:
                objeto_contacto.horario = contacto['horario'].title()

            objeto_contacto.save()

            log("Se ha creado el registro:")
            log("Nombre: ", objeto_contacto, "\n Teléfono Particular ", objeto_contacto.telefono_particular, "\n Teléfono Celular: ", objeto_contacto.telefono_celular, "\n Email: ", objeto_contacto.email, "\n Horario: ", objeto_contacto.horario)
            log("===========")

    def importar_pisos(self):
        pisos = self.obtener_datos_desde_api('pisos')['pisos']

        print("Importando Pisos")
        bar = barra_de_progreso(simple=False)

        for piso in bar(pisos):

            if piso['marca']:
                marca = piso['marca']
            else:
                marca = "Desconocido"

            log("Buscando piso para escuela: ", piso['cue'])
            objeto_escuela = models.Escuela.objects.get(cue=piso['cue'])
            objeto_piso, created = models.Piso.objects.get_or_create(legacy_id=piso['legacy_id'])
            #
            objeto_piso.servidor = piso['marca']
            objeto_piso.serie = piso['serie']

            if piso['ups']:
                if piso['ups'] == "SI":
                    objeto_piso.ups = True
                else:
                    objeto_piso.ups = False
            else:
                objeto_piso.ups = False


            if piso['rack']:
                if piso['rack'] == "SI":
                    objeto_piso.rack = True
                else:
                    objeto_piso.rack = False
            else:
                objeto_piso.rack = False

            objeto_piso.estado = piso['piso_estado']

            objeto_escuela.piso = objeto_piso

            objeto_piso.save()
            objeto_escuela.save()

            log("Se ha creado el registro:")
            log("Piso de escuela ", piso['cue'], ": \n Servidor: ", piso['marca'], "\n Serie: ", piso['serie'], "\n UPS: ", piso['ups'], "\n Rack: ", piso['rack'], "\n Estado: ", piso['piso_estado'])
            log("===========")

    def importar_tareas(self):
        tareas = self.obtener_datos_desde_api('tickets')['tickets']
        cantidad_de_tareas_creadas = 0
        cantidad_de_tareas_omitidas = 0

        print("Importando Tareas")
        bar = barra_de_progreso(simple=False)

        for tarea in bar(tareas):
            log("Se intenta crear el registro con id_original: " + str(tarea['id_ticket_original']) + " y DNI de usuario: " + str(tarea['dni_usuario']))
            log("Prioridad de la tarea: " + str(tarea['prioridad']))
            log("Motivo de la tarea: " + unicode(tarea['motivo']))
            log("Estado de la tarea: " + str(tarea['estado']))

            dni_usuario = tarea['dni_usuario']

            try:
                objeto_autor = models.Perfil.objects.get(dni=dni_usuario)
            except models.Perfil.DoesNotExist:
                log("Error, no existe registro de usuario buscado %s. No se registrará la tarea." %(dni_usuario))
                cantidad_de_tareas_omitidas += 1
                continue

            objeto_tarea, created = models.Tarea.objects.get_or_create(id_ticket_original=tarea['id_ticket_original'])

            objeto_escuela = models.Escuela.objects.get(cue=tarea['cue'])
            objeto_motivo = models.MotivoDeTarea.objects.get(nombre=tarea['motivo'])
            objeto_estado = models.EstadoDeTarea.objects.get(nombre=tarea['estado'])

            prioridad = tarea['prioridad']
            if prioridad == 1:
                prioridad = "Alta"
            elif prioridad == 2:
                prioridad = "Media"
            elif prioridad == 3:
                prioridad = "Baja"

            objeto_prioridad = models.PrioridadDeTarea.objects.get(nombre=prioridad)

            fecha_alta = tarea['fecha_alta']
            #
            objeto_tarea.fecha_de_alta = fecha_alta
            objeto_tarea.titulo = "Tarea #: " + str(tarea['id_ticket_original'])
            objeto_tarea.descripcion = tarea['descripcion']
            objeto_tarea.autor = objeto_autor
            objeto_tarea.escuela = objeto_escuela
            objeto_tarea.motivo_de_tarea = objeto_motivo
            objeto_tarea.estado_de_tarea = objeto_estado
            objeto_tarea.prioridad_de_tarea = objeto_prioridad

            objeto_tarea.save()


            log("Se ha creado el registro:")
            log("Tarea con id_original: " + str(tarea['id_ticket_original']))
            log("===========")
            cantidad_de_tareas_creadas += 1

        print("Resumen de tareas:")
        print("   Se crearon %d tareas correctamente." %(cantidad_de_tareas_creadas))
        print("   Se evitaron crear %d tareas porque correspondían a usuarios inexistentes." %(cantidad_de_tareas_omitidas))

    def importar_comentarios_de_tareas(self):
        comentarios = self.obtener_datos_desde_api('comentarios_tickets')['comentarios_tickets']
        cantidad_de_comentarios_de_tareas_creados = 0
        cantidad_de_comentarios_de_tareas_omitidos = 0

        print("Importando Comentarios de Tareas")
        bar = barra_de_progreso(simple=False)

        for comentario in bar(comentarios):
            log("Se intenta crear el registro con id_original: " + str(comentario['id_ticket_original']) + " y DNI de usuario: " + str(comentario['dni_usuario']))

            dni_usuario = comentario['dni_usuario']
            id_ticket_original = comentario['id_ticket_original']

            try:
                objeto_autor = models.Perfil.objects.get(dni=dni_usuario)
            except models.Perfil.DoesNotExist:
                log("Error, no existe registro de usuario buscado %s. No se registrará la tarea." %(dni_usuario))
                cantidad_de_comentarios_de_tareas_omitidos += 1
                continue

            try:
                objeto_tarea = models.Tarea.objects.get(id_ticket_original=id_ticket_original)
            except models.Tarea.DoesNotExist:
                log("Error, no existe registro de tarea buscado %s. No se registrará el comentario." %(id_ticket_original))
                cantidad_de_comentarios_de_tareas_omitidos += 1
                continue



            objeto_comentario, created = models.ComentarioDeTarea.objects.get_or_create(comentario=comentario['comentario'])
            objeto_comentario.autor = objeto_autor
            objeto_comentario.fechaDeAlta = comentario['fecha']
            objeto_comentario.tarea = objeto_tarea

            objeto_comentario.save()


            log("Se ha creado el registro:")
            log("Comentario de Tarea con id_original: " + str(comentario['id_ticket_original']))
            log("===========")
            cantidad_de_comentarios_de_tareas_creados += 1

        print("Resumen de tareas:")
        print("   Se crearon %d comentarios de tareas correctamente." %(cantidad_de_comentarios_de_tareas_creados))
        print("   Se evitaron crear %d comentarios de tareas porque correspondían a usuarios inexistentes." %(cantidad_de_comentarios_de_tareas_omitidos))

    def vincular_programas(self):
        programas = self.obtener_datos_desde_api('programas')['programas']

        print("Vinculando Programas")
        bar = barra_de_progreso(simple=False)

        for programa in bar(programas):
            log("Busando programas para escuela: ", programa['cue'])

            objeto_escuela = models.Escuela.objects.get(cue=programa['cue'])

            objeto_escuela.programas.add(models.Programa.objects.get(nombre=programa['programa']))

            objeto_escuela.save()

            log("Se ha vinculado el registro:")
            log("Programa: ", programa['programa'], "a la escuela con CUE ", programa['cue'])
            log("===========")

    def vincular_acompaniantes(self):
        acompaniantes = self.obtener_datos_desde_api('acompaniantes_eventos')['acompaniantes_eventos']

        print("Vinculando Acompañantes de eventos")
        bar = barra_de_progreso(simple=False)

        for acompaniante in bar(acompaniantes):

            legacy_id = acompaniante['legacy_id']
            dni_usuario = acompaniante['dni_usuario']

            log("Busando acompaniantes para legacy_id: ", legacy_id)

            try:
                objeto_evento = models.Evento.objects.get(legacy_id=legacy_id)
            except models.Evento.DoesNotExist:
                log("Error, no existe registro de evento con legacy_id %s. No se registrará el acompañante." %(legacy_id))
                # cantidad_de_comentarios_de_tareas_omitidos += 1
                continue

            try:
                models.Perfil.objects.get(dni=dni_usuario)
            except models.Perfil.DoesNotExist:
                log("Error, no existe registro de usuario buscado %s. No se registrará el acompañante." %(dni_usuario))
                # cantidad_de_tareas_omitidas += 1
                continue

            objeto_evento.acompaniantes.add(models.Perfil.objects.get(dni=dni_usuario))

            objeto_evento.save()

            log("Se ha vinculado el registro:")
            log("Acomaniante: ", acompaniante['nombre'], "al evento con legacy_id ", acompaniante['legacy_id'])
            log("===========")

    def vincular_acompaniantes_por_perfil(self):
        if PERFIL:
            print (u"El ID de perfil es " + PERFIL)
        ruta = "acompaniantes_eventos_por_perfil?perfil_id=" + PERFIL
        acompaniantes = self.obtener_datos_desde_api(ruta)['acompaniantes_eventos_por_perfil']

        print("Vinculando Acompañantes de eventos")
        bar = barra_de_progreso(simple=False)

        for acompaniante in bar(acompaniantes):

            legacy_id = acompaniante['legacy_id']
            dni_usuario = acompaniante['dni_usuario']

            log("Busando acompaniantes para legacy_id: ", str(legacy_id))

            try:
                objeto_evento = models.Evento.objects.get(legacy_id=legacy_id)
            except models.Evento.DoesNotExist:
                log("Error, no existe registro de evento con legacy_id %s. No se registrará el acompañante." %(legacy_id))
                # cantidad_de_comentarios_de_tareas_omitidos += 1
                continue

            try:
                models.Perfil.objects.get(dni=dni_usuario)
            except models.Perfil.DoesNotExist:
                log("Error, no existe registro de usuario buscado %s. No se registrará el acompañante." %(dni_usuario))
                # cantidad_de_tareas_omitidas += 1
                continue

            objeto_evento.acompaniantes.add(models.Perfil.objects.get(dni=dni_usuario))

            objeto_evento.save()

            log("Se ha vinculado el registro:")
            log("Acomaniante: ", acompaniante['nombre'], "al evento con legacy_id ", str(acompaniante['legacy_id']))
            log("===========")

    def obtener_datos_desde_api(self, data):
        url = BASE_URL + data
        print("Consultando la URL: " + url)
        resultado = requests.get(url)
        return resultado.json()

    def crear_tipos_de_financiamiento(self):
        nombres = ["Nacional", "Provincial", "Municipal", "Propio"]

        print("Creando Tipos de Financiamiento")
        bar = barra_de_progreso()

        for nombre in bar(nombres):
            p, created = models.TipoDeFinanciamiento.objects.get_or_create(nombre=nombre)
            log(str(p))

    def crear_estados_de_paquetes(self):
        nombres = ["Objetado", "Pendiente", "EducAr", "Devuelto", "Descargado"]

        print("Creando Estados de Paquetes")
        bar = barra_de_progreso()

        for nombre in bar(nombres):
            p, created = models.EstadoDePaquete.objects.get_or_create(nombre=nombre)
            log(str(p))

    def crear_motivos_de_tareas(self):
        nombres = [
            "Servidor robado",
            "Servidor roto",
            "Piso tecnológico",
            "Paquetes de provisión",
            "Movimiento de equipamiento",
            "Problemas eléctricos",
            "Switch roto",
            "UPS roto",
            "Mantenimiento básico de piso",
            "Ampliacion de piso",
            "Reingeniería de piso",
            "Mudanza de piso",
            "Reclamos del territorio"
        ]

        print("Creando Motivos de Tareas")
        bar = barra_de_progreso()

        for nombre in bar(nombres):
            p, created = models.MotivoDeTarea.objects.get_or_create(nombre=nombre)
            log(str(p))

    def crear_motivos_de_conformaciones(self):
        nombres = [
            "Comparte Piso",
            "Comparte Edificio",
            "CUE Nuevo",
            "CUE Anterior",
            "Se Unificó"
        ]

        print("Creando Motivos de Conformaciones")
        bar = barra_de_progreso()

        for nombre in bar(nombres):
            p, created = models.MotivoDeConformacion.objects.get_or_create(nombre=nombre)
            log(str(p))

    def crear_estados_de_tareas(self):
        nombres = [
            "Abierto",
            "En Progreso",
            "En Espera",
            "Cerrado"
        ]

        print("Creando Estados de Tareas")
        bar = barra_de_progreso()

        for nombre in bar(nombres):
            p, created = models.EstadoDeTarea.objects.get_or_create(nombre=nombre)
            log(str(p))

    def crear_estados_de_validaciones(self):
        nombres = [
            "Pendiente",
            "Objetada",
            "Aprobada"
            # "Pendiente", #1
            # "Revisión", #2
            # "Cerrado", #3,
            # "Eliminado", #4
            # "Conformacion", #5
            # "No valida" #6
        ]

        print("Creando Estados de Validaciones")
        bar = barra_de_progreso()

        for nombre in bar(nombres):
            p, created = models.EstadoDeValidacion.objects.get_or_create(nombre=nombre)
            log(str(p))

    def crear_prioridades_de_tareas(self):
        nombres = [
            "Alta",
            "Media",
            "Baja"
        ]

        print("Creando Prioridades de Tareas")
        bar = barra_de_progreso()

        for nombre in bar(nombres):
            p, created = models.PrioridadDeTarea.objects.get_or_create(nombre=nombre)
            log(str(p))

    def crear_niveles(self):
        nombres = ["Inicial", "Primaria", "Secundaria", "Superior"]

        print("Creando Niveles")
        bar = barra_de_progreso()

        for nombre in bar(nombres):
            p, created = models.Nivel.objects.get_or_create(nombre=nombre)
            log(str(p))

    def crear_modalidades(self):
        nombres = ["Técnica", "Especial", "Ninguna", "Artística"]

        print("Creando Modalidades")
        bar = barra_de_progreso()

        for nombre in bar(nombres):
            p, created = models.Modalidad.objects.get_or_create(nombre=nombre)
            log(str(p))

    def crear_tipos_de_gestion(self):
        nombres = ["Estatal", "Privada", "Compartida"]

        print("Creando Tipos de Gestión")
        bar = barra_de_progreso()

        for nombre in bar(nombres):
            p, created = models.TipoDeGestion.objects.get_or_create(nombre=nombre)
            log(str(p))

    def crear_areas(self):
        nombres = ["Urbana", "Rural"]

        print("Creando Areas")
        bar = barra_de_progreso()

        for nombre in bar(nombres):
            p, created = models.Area.objects.get_or_create(nombre=nombre)
            log(str(p))

    def crear_programas(self):
        nombres = [
            "Conectar Igualdad",
            "PAD",
            "Responsabilidad Empresarial",
            "Primaria Digital",
            "Escuelas del Futuro"
            ]

        print("Creando Programas")
        bar = barra_de_progreso()

        for nombre in bar(nombres):
            p, created = models.Programa.objects.get_or_create(nombre=nombre)
            log(str(p))

    def crear_experiencias(self):
        nombres = [
            "Técnico",
            "Pedagógico",
            "Administrativo",
            "Diseño",
            "Comunicación"
            ]

        print("Creando Experiencias")
        bar = barra_de_progreso()

        for nombre in bar(nombres):
            p, created = models.Experiencia.objects.get_or_create(nombre=nombre)
            log(str(p))

    def crear_contratos(self):
        nombres = [
            "PLANIED",
            "Planta/PLANIED",
            "Planta",
            "ConIg",
            "Ord. Tec.",
            "PLANIED/EDF"
        ]

        print("Creando Contactos")
        bar = barra_de_progreso()

        for nombre in bar(nombres):
            p, created = models.Contrato.objects.get_or_create(nombre=nombre)
            log(str(p), str(created))

    def crear_cargos(self):
        nombres = [
            ("FED", "Facilitador Educación Digital"),
            ("FEF", "FEF"),
            ("Coord", "Coordinador"),
            ("Adm", "Administrativo"),
            ("Coord EF", "Coordinador EF"),
            ("FED esp", "Facilitador Educación Digital Especial"),
            ("Coord Prov", "Coordinador Provincial")
            ]

        print("Creando Cargos")
        bar = barra_de_progreso()

        for nombre in bar(nombres):
            p, created = models.Cargo.objects.get_or_create(nombre=nombre[0], descripcion=nombre[1])
            log(str(p))

    def crear_cargos_escolares(self):
        nombres = [
            "Director",
            "Vice Director",
            "Secretario",
            "Maestro",
            "EMATP",
            "Prosecretario",
            "Preceptor",
            "Profesor",
            "Otro"
            ]

        print("Creando Cargos Escolares")
        bar = barra_de_progreso()

        for nombre in bar(nombres):
            p, created = models.CargoEscolar.objects.get_or_create(nombre=nombre)
            log(str(p))

    def crear_categorias_de_eventos(self):
        nombres = [
            "Acciones especiales/Congresos",
            "Acciones especiales/Desembarcos",
            "Acciones especiales/Encuentros masivos",
            "Acciones especiales/Prácticas profesionales",
            "Acciones especiales/Otros",
            "Asistencia/Administrativa",
            "Asistencia/Pedagógica",
            "Asistencia/Técnica",
            "Escuelas del futuro/Seguimiento de actividades",
            "Escuelas del futuro/Seguimiento de acciones",
            "Capacitaciones/Sensibilización",
            "Capacitaciones/Docentes",
            "Capacitaciones/Alumnos",
            "Capacitaciones/Capacitación",
            "Reunión/Online",
            "Reunión/Inspectores",
            "Reunión/Referente de área",
            "Reunión/Equipo",
            "Reunión/Planificación",
            "Reunión/Región Central",
            "Visita pedagógica/Comunicación",
            "Visita pedagógica/Docentes",
            "Visita pedagógica/Educación especial",
            "Visita pedagógica/Equipos directivos",
            "Visita pedagógica/Equipos jurisdiccionales",
            "Visita pedagógica/Equipos de gestión tic",
            "Visita pedagógica/Evaluación y seguimiento",
            "Visita pedagógica/Jóvenes",
            "Visita pedagógica/Otras",
            "Visita pedagógica/Primera entrega",
            "Visita pedagógica/Rte",
            "Visita pedagógica/Supervisores y directores de nivel y modalidad",
            "Visita técnica/Carga de matrícula",
            "Visita técnica/Configuración y registro de netbook",
            "Visita técnica/Desbloqueo de netbook",
            "Visita técnica/Mantenimiento de piso tecnológico",
            "Visita técnica/Otras",
            "Visita técnica/Primera entrega",
            "Visita técnica/Servicio técnico a netbook"
            ]

        print("Creando Categorías de eventos")
        bar = barra_de_progreso()

        for nombre in bar(nombres):
            p, created = models.CategoriaDeEvento.objects.get_or_create(nombre=nombre)
            log(str(p))

    def limpiar_e_importar_permisos_con_grupos(self):
        # Elimina todos los permisos por omision de django.
        print("Borrando los permisos estándar de django")
        permisos_estandar = [x for x in Permission.objects.all() if x.name.startswith('Can ')]

        if permisos_estandar:
            print("Borrando %d" %(len(permisos_estandar)))
            [x.delete() for x in permisos_estandar]
        else:
            print("No hay permisos estandar para borrar")

        print("Borrando el resto de los permisos para re-definir")
        [x.delete() for x in Permission.objects.all()]

        # Genera los permisos personalizados

        # Acceso a todas las regiones.
        PERFIL_GLOBAL = 'perfil.global'

        AGENDA_LISTAR = 'agenda.listar'
        AGENDA_CREAR = 'agenda.crear'
        AGENDA_COMENTAR = 'agenda.comentar'

        TAREAS_LISTAR = 'tareas.listar'
        TAREAS_CREAR = 'tareas.crear'
        TAREAS_EDITAR = 'tareas.editar'
        TAREAS_COMENTAR = 'tareas.comentar'
        TAREAS_CERRAR = 'tareas.cerrar'

        ESCUELAS_LISTAR = 'escuelas.listar'
        ESCUELAS_EDITAR = 'escuelas.editar'
        ESCUELAS_CONFORMAR = 'escuelas.conformar'
        ESCUELAS_CAMBIAR_ESTADO = 'escuelas.cambiarestado'
        ESCUELAS_ELIMINAR = 'escuelas.eliminar'
        ESCUELAS_CREAR = 'escuelas.crear'
        ESCUELAS_VER_MAPA = 'escuelas.vermapa'

        VALIDACIONES_LISTAR = 'validaciones.listar'
        VALIDACIONES_CREAR = 'validaciones.crear'
        VALIDACIONES_COMENTAR = 'validaciones.comentar'
        VALIDACIONES_EDITAR = 'validaciones.editar'

        ENTREGAS_LISTAR = 'entregas.listar'
        ENTREGAS_ACTUALIZAR = 'entregas.actualizar'

        PAQUETES_LISTAR = 'paquetes.listar'
        PAQUETES_CREAR = 'paquetes.crear'
        PAQUETES_EDITAR = 'paquetes.editar'
        PAQUETES_EXPORTAR = 'paquetes.exportar'

        PERSONAS_LISTAR = 'personas.listar'
        PERSONAS_CREAR = 'personas.crear'
        PERSONAS_EDITAR = 'personas.editar'
        PERSONAS_ELIMIAR = 'personas.eliminar'
        PERSONAS_CAMBIAR_ESTADO = 'personas.cambiarestado'
        PERSONAS_VER_INFORME = 'personas.verinformes'

        MATRIX_LISTAR = 'matrix.listar'

        permisos = [
            PERFIL_GLOBAL,
            AGENDA_LISTAR, AGENDA_CREAR, AGENDA_COMENTAR,
            TAREAS_LISTAR, TAREAS_CREAR, TAREAS_EDITAR, TAREAS_COMENTAR, TAREAS_CERRAR,
            ESCUELAS_LISTAR, ESCUELAS_EDITAR, ESCUELAS_CONFORMAR, ESCUELAS_CAMBIAR_ESTADO, ESCUELAS_ELIMINAR, ESCUELAS_CREAR, ESCUELAS_VER_MAPA,
            VALIDACIONES_LISTAR, VALIDACIONES_CREAR, VALIDACIONES_COMENTAR, VALIDACIONES_EDITAR,
            ENTREGAS_LISTAR, ENTREGAS_ACTUALIZAR,
            PAQUETES_LISTAR, PAQUETES_CREAR, PAQUETES_EDITAR, PAQUETES_EXPORTAR,
            PERSONAS_LISTAR, PERSONAS_CREAR, PERSONAS_EDITAR, PERSONAS_ELIMIAR, PERSONAS_CAMBIAR_ESTADO, PERSONAS_VER_INFORME,
            MATRIX_LISTAR
        ]

        print("Actualizando el listado de permisos (creación o actualización)")
        bar = barra_de_progreso()

        for p in bar(permisos):
            modelo, permiso = p.split('.')

            tipo, _ = ContentType.objects.get_or_create(app_label='escuelas', model=modelo)
            Permission.objects.get_or_create(name=permiso, codename=p, content_type=tipo)

        grupos = {
            'Coordinador': [
                ESCUELAS_LISTAR, ESCUELAS_CONFORMAR, ESCUELAS_EDITAR,
                AGENDA_LISTAR, AGENDA_CREAR, AGENDA_COMENTAR,
                TAREAS_LISTAR, TAREAS_CREAR, TAREAS_COMENTAR, TAREAS_CERRAR,
                VALIDACIONES_CREAR, VALIDACIONES_LISTAR, VALIDACIONES_COMENTAR, VALIDACIONES_EDITAR,
                ENTREGAS_ACTUALIZAR,
                PAQUETES_LISTAR, PAQUETES_CREAR, PAQUETES_EDITAR,
                PERSONAS_LISTAR,
            ],
            'Invitado': [
                ESCUELAS_LISTAR,
                AGENDA_LISTAR,
            ],
            'Sin Definir': [
                ESCUELAS_LISTAR,
            ],
            'Administrador': permisos,
            'Facilitador': [
                ESCUELAS_LISTAR, ESCUELAS_EDITAR,
                AGENDA_LISTAR, AGENDA_CREAR, AGENDA_COMENTAR,
                TAREAS_LISTAR, TAREAS_CREAR, TAREAS_COMENTAR, TAREAS_CERRAR,
                PAQUETES_LISTAR, PAQUETES_CREAR,
                PERSONAS_LISTAR,
            ],
            'Referente': [
                PERFIL_GLOBAL,
                ESCUELAS_LISTAR,
                AGENDA_LISTAR, AGENDA_CREAR, AGENDA_COMENTAR,
                TAREAS_LISTAR, TAREAS_CREAR, TAREAS_COMENTAR, TAREAS_CERRAR,
                ESCUELAS_LISTAR,
                PERSONAS_LISTAR,
            ],
            'Administración': [
                AGENDA_LISTAR,
                TAREAS_LISTAR, TAREAS_CREAR, TAREAS_COMENTAR, TAREAS_CERRAR,
                ESCUELAS_LISTAR,
                PERSONAS_LISTAR, PERSONAS_CREAR, PERSONAS_EDITAR, PERSONAS_ELIMIAR, PERSONAS_CAMBIAR_ESTADO, PERSONAS_VER_INFORME,
                PERFIL_GLOBAL,
            ]
        }

        for nombre_de_grupo in grupos:

            (grupo, _) = Group.objects.get_or_create(name=nombre_de_grupo)

            for nombre_de_permiso in grupos[nombre_de_grupo]:
                permiso = Permission.objects.get(codename=nombre_de_permiso)
                grupo.permissions.add(permiso)


        """
        print("Realizando asignación de grupos")

        asignaciones = [
            # ( Email laboral  ,  Nombre del grupo )
            ('ccane@abc.gob.ar', 'Administrador'),
            ('lvigolo@abc.gob.ar', 'Administrador'),
        ]

        for emailLaboral, grupo in asignaciones:
            try:
                perfil = models.Perfil.objects.get(emailLaboral=emailLaboral)
            except models.Perfil.DoesNotExist:
                raise TypeError("Hay un usuario duplicado " + emailLaboral)
            perfil.definir_grupo_usando_nombre(grupo)
            perfil.save()
        """

    def aplicar_permiso_sin_definir_a_los_perfiles_faltantes(self):

        print("Aplicando el grupo 'Sin definir' a todos los perfiles que no tengan grupo")
        bar = barra_de_progreso()

        for perfil in bar(models.Perfil.objects.all()):
            if not perfil.group:
                perfil.group = Group.objects.get(name='Sin Definir')
                perfil.save()

    def aplicar_recibido_a_paquetes(self):

        print("Asignando estado Recibido a paquetes de Enero de 2018")
        bar = barra_de_progreso()

        inicio = "2018-01-01"
        fin = "2018-01-31"

        paquetes_modificados = 0

        total_paquetes = models.Paquete.objects.all()
        paquetes_2017 = total_paquetes.filter(fecha_pedido__range=(inicio, fin))
        paquetes_2017_objetados = paquetes_2017.filter(estado__nombre="Objetado").distinct()
        paquetes_2017_pendientes = paquetes_2017.filter(estado__nombre="Pendiente").distinct()
        paquetes_2017_enviados = paquetes_2017.filter(estado__nombre="EducAr").distinct()
        print("Total de paquetes: " + str(total_paquetes.count()))
        print("Paquetes Enero 2018: " + str(paquetes_2017.count()))
        print("Paquetes Enero 2018 Objetados: " + str(paquetes_2017_objetados.count()))
        print("Paquetes Enero 2018 Pendientes: " + str(paquetes_2017_pendientes.count()))
        print("Paquetes Enero 2018 Enviados: " + str(paquetes_2017_pendientes.count()))

        estado_recibido = models.EstadoDePaquete.objects.get(nombre="Devuelto")

        for paquete in paquetes_2017_pendientes:
            print("Se va a cambiar el estado del paquete id " + str(paquete.id) + " a Devuelto...")
            paquete.estado = estado_recibido
            paquete.save()
            print("Estado cambiado correctamente")
            paquetes_modificados += 1

        print ("Finalizó la actualización de estados de paquetes Enero 2018.")
        print ("Se modificaron " + str(paquetes_modificados) + " paquetes.")



    def importar_estado_de_paquetes(self):
        # ARCHIVO = './/archivos_para_importacion/dte_perfiles_2017.xlsx'
        ARCHIVO = './/archivos_para_importacion/objetados.xlsx'
        LIMITE_DE_FILAS = 800

        print("Comenzando la importación de paquetes")
        log("Iniciando la importación del archivo: " + ARCHIVO)
        wb = load_workbook(ARCHIVO)

        columnas_como_string = ", ".join(wb.get_sheet_names())
        log("Las páginas de la planilla son: " + columnas_como_string)

        filas_procesadas = 0
        filas_omitidas_o_con_errores = 0
        filas_omitidas_lista = ""

        listado = ""

        def formatear_fecha(fecha):
            if fecha:
                return fecha.strftime('%Y-%m-%d')
            else:
                return fecha

        def obtener_valores_desde_fila(fila):
            return {
                "cue":                  unicode(fila[0].value),
                "numero_de_servidor":   unicode(fila[1].value),
                "ne":                   unicode(fila[2].value),
                "hardware_id":          unicode(fila[3].value),
                "marca_de_arranque":    unicode(fila[4].value),
                "motivo_de_objecion":   unicode(fila[5].value),
            }

        bar = barra_de_progreso(simple=False)
        #for conformacion in bar(conformaciones):

        cantidad_de_filas_con_datos = 0
        cantidad_de_filas_procesadas_sin_errores = 0

        for indice, fila in bar(enumerate(wb.active.rows)):

            if indice is 0:
                continue;             # Ignora la cabecera

            if not fila[0].value: # Se elige la primer columna porque siempre va a haber un CUE, pero puede no haber número de servidor.
                log("Terminando en la fila %d porque no parece haber mas registros." %(indice + 1))
                break

            cantidad_de_filas_con_datos += 1
            log("Procesando fila '%d'" %(indice +1))

            try:
                valores = obtener_valores_desde_fila(fila)

                cue = valores['cue']
                numero_de_servidor = valores['numero_de_servidor']
                ne = valores['ne']
                hardware_id = valores['hardware_id']
                marca_de_arranque = valores['marca_de_arranque']
                motivo_de_objecion = valores['motivo_de_objecion']

                print("-----------------------------------------")
                print("Se va a procesar el paquete:")
                print("-----------------------------------------")
                print("CUE: " + cue)
                print("Nro de servidor: " + numero_de_servidor)
                print("N/E: " + ne)
                print("HWID: " + hardware_id)
                print("Marca de arranque: " + marca_de_arranque)
                print("Motivo: " + motivo_de_objecion)

                estado_objetado = models.EstadoDePaquete.objects.get(nombre="Objetado")

                try:
                    paquete = models.Paquete.objects.get(escuela__cue=cue, ne=ne, id_hardware=hardware_id)
                except models.Paquete.DoesNotExist:
                    print("-----------------------------------")
                    print("No existe un paquete con esos datos")
                    print("-----------------------------------")
                    listado += cue + ";" + numero_de_servidor + ";" + ne + ";" + hardware_id + ";" + marca_de_arranque + ";" + motivo_de_objecion + ";Falló porque no fue encontrado. \n"
                    continue
                except models.Paquete.MultipleObjectsReturned:
                    print("-----------------------------------")
                    print("Mas de un registro encontrado")
                    print("-----------------------------------")
                    listado += cue + ";" + numero_de_servidor + ";" + ne + ";" + hardware_id + ";" + marca_de_arranque + ";" + motivo_de_objecion + ";Falló porque se encontró mas de un registro.\n"
                    continue


                paquete.estado = estado_objetado
                paquete.comentario = motivo_de_objecion
                paquete.save()

                cantidad_de_filas_procesadas_sin_errores += 1

            except TypeError, e:
                log("-----")
                log("Fila %d - ****** OMITIDA, TypeError. La fila contiene caracteres incorrectos." %(indice + 1))
                filas_omitidas_o_con_errores += 1
                filas_omitidas_lista += ", " + str(indice + 1)
                log(str(e))
                log("-----")
                continue

            # log("Fila %d - Cargando datos de perfil para consultor: '%s'" %(indice + 1, valores["consultor"]))

            # listado += apellido + nombre + "," + region + "," + email_laboral + "," + default_pass + "\n"


            filas_procesadas += 1

            if indice > LIMITE_DE_FILAS:
                break


        log("Terminó la ejecución")

        print("")
        print("Resumen:")
        print("")
        print(" - cantidad total de filas:                       " + str(cantidad_de_filas_con_datos))
        print(" - filas procesadas:                              " + str(cantidad_de_filas_procesadas_sin_errores))
        print(" - cantidad de filas que fallaron:                " + str(cantidad_de_filas_con_datos - cantidad_de_filas_procesadas_sin_errores))

        print("")
        print("")
        print("Listado: ")
        print("")
        print listado
        f = open('listado_de_paquetes_fallados.csv', 'w')
        f.write(listado.encode('utf-8'))
        f.close()
